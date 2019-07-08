#Manipulates data according to what PRICE_UPDATES equals
#Can also apply 10% discount across all produces

import openpyxl

#1. Read Excel sheet
print("Opening Workbook...")
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

#The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

#Expirment: discount = 1 - 0.9 = 10% discount
#discount = 0.9


#2. Loop through the rows and update the prices.
for rowNUM in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row=rowNUM, column=1).value
    #sheet.cell(row=rowNUM, column=2).value *= discount

    #If produce name appears in PRICE_UPDATES update price
    if produceName in PRICE_UPDATES:
       sheet.cell(row=rowNUM, column=2).value = PRICE_UPDATES[produceName]

#3. Save to different Excel file
wb.save('updatedProduceSales.xlsx')
